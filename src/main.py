import os
import smtplib
import sys
import webbrowser
from openpyxl import load_workbook
from email.message import EmailMessage

import pytoml

# Final Variables
CONFIG_PATH = 'config.toml'


# Function to import from config.toml, which contains environment variables
def read_config(path):
    config = pytoml.load(open(path, 'rb'))
    return config


# Function to create connection to localhost for testing purposes
# For windows: open cmd and input <py -m smtpd -c DebuggingServer -n localhost:1025>
# For MacOS: open terminal and input <python3 -m smtpd -c DebuggingServer -n localhost:1025>
# For testing purpose, EMAIL_ADDRESS AND receiver can be blank strings
def debug_local_server(subject, body):
    test_email = ""
    receiver = ""
    with smtplib.SMTP('localhost', 1025) as smtp:
        message = f'Subject: {subject}\n\n{body}'
        smtp.sendmail(test_email, receiver, message)


# Function to change directories
def change_directory(current, extension):
    os.chdir(current + extension)


# Function to list files in a directory
def list_files(path):
    files = os.listdir(path)
    return files


# Function to read attachments before attaching to email message
def read_files(lst):
    for file in lst:
        with open(file, 'rb') as f:
            file_data = f.read()
            file_name = f.name
    return file_data, file_name


# Function to get email_receivers and relevant content from email_list directory
def get_receivers_info(xlsx_wb):
    email_list_worksheet = xlsx_wb['Sheet1']
    data = []
    for rows in email_list_worksheet.iter_rows(min_row=2):
        data.append((rows[0].value, rows[1].value))
    return data


# TODO: Work on function to print email body in python console for user to check through before sending
def save_to_html(html_str):
    file = open("template.html", "w")
    file.write(html_str)
    file.close()


if __name__ == '__main__':
    # Read config.toml
    config_file = read_config('config.toml')

    # Store cwd for future use
    base_cwd = os.getcwd()

    # Obtain login details from config.toml
    email_address = config_file['sender']['email_address']
    email_password = config_file['sender']['password']
    email_server = config_file['sender']['server']

    # Navigate to email_list directory
    email_list_directory = config_file["directories"]["email_list"]
    change_directory(base_cwd, email_list_directory)
    email_list_files = list_files(os.getcwd())

    # Read "email_list.xlsx" file
    # TODO: Add try-catch for absence of "email_list.xlsx
    email_list_xlsx = load_workbook(filename="email_list.xlsx")
    email_receiver_list = get_receivers_info(email_list_xlsx)

    # Start of for loop to loop through email_receiver_list
    for receiver in email_receiver_list:
        replacements = {'organisation_name': receiver[0]}
        # Draft message
        msg = EmailMessage()
        msg["Subject"] = "Stop lying pls"
        msg["From"] = email_address
        msg["To"] = receiver[1]

        # Set short_name and full_name variable from config.toml file
        # TODO: Add try-catch to handle case where variables are wrong (hort_name and full_name changed in config.toml)
        for var, name in config_file['email_content'].items():
            replacements[var] = name

        # Navigate to email_script directory
        email_script_directory = config_file["directories"]["email_script"]
        change_directory(base_cwd, email_script_directory)
        email_template = list_files(os.getcwd())  # List files in email_script directory

        # TODO: Add try-catch to handle case where "template.txt" doesn't exist
        if len(email_template) == 2 and "template.txt" in email_template:  # Check for existence of "template.txt"
            with open(email_template[email_template.index("template.txt")], mode='r', encoding='utf-8-sig') as fle:
                html_string = fle.read().format(**replacements)  # read template.txt as string

        msg.set_content(html_string, subtype='html')  # convert to html and attach to msg object
        save_to_html(html_string)  # update to "template.html"
        webbrowser.get('windows-default').open_new("template.html")  # open in default browser
        # Request for user to check email
        check_preview = input(f"Please check preview of email to {msg['To']} and choose whether to proceed (YES/NO): ")
        if check_preview != "YES":
            print("Stopping program...")
            sys.exit()

        # Navigate to attachment directory
        attachment_directory = config_file["directories"]["attachments"]
        change_directory(base_cwd, attachment_directory)

        # List attachments in attachment directory
        attachments = list_files(os.getcwd())

        # Attach attachments to msg
        f_data, f_name = read_files(attachments)

        # Uncomment when reading to send
        msg.add_attachment(f_data, maintype='application', subtype='octet-stream', filename=f_name)
        # Uncomment when ready to send email
        with smtplib.SMTP_SSL(email_server, 465) as smtp:
            smtp.login(email_address, email_password)
            smtp.send_message(msg)
            print(f'Email successfully sent to {msg["To"]}.')
